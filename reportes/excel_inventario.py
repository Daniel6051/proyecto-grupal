from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def exportar_excel_inventario(inventario, ruta_salida="inventario.xlsx"):
    """
    Exporta el estado actual del inventario a Excel con formato.
    'inventario' puede ser lista de diccionarios o de objetos Producto.
    """
    os.makedirs(os.path.dirname(ruta_salida) if os.path.dirname(ruta_salida) else '.', exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Actual"

    # ── Estilos ────────────────────────────────────────────────────────
    color_header    = "1e1e2e"   # fondo oscuro encabezado
    color_texto_hdr = "cba6f7"   # violeta para el texto del header
    color_alerta    = "f38ba8"   # rojo suave para stock crítico
    color_ok        = "a6e3a1"   # verde suave para stock normal
    color_fila_alt  = "313244"   # gris oscuro para filas alternas

    fuente_header = Font(name="Calibri", bold=True, color=color_texto_hdr, size=11)
    fuente_normal = Font(name="Calibri", color="cdd6f4", size=10)
    relleno_header = PatternFill("solid", fgColor=color_header)
    borde = Border(
        bottom=Side(style="thin", color="6c7086"),
        right=Side(style="thin", color="6c7086"),
    )

    # ── Título ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    titulo = ws["A1"]
    titulo.value = "Reporte de Inventario — Stock Actual"
    titulo.font = Font(name="Calibri", bold=True, size=14, color="cba6f7")
    titulo.fill = PatternFill("solid", fgColor="181825")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Encabezados ────────────────────────────────────────────────────
    encabezados = ["ID", "Nombre del Producto", "Stock Actual", "Stock Mínimo", "Precio ($)"]
    for col_idx, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=2, column=col_idx, value=texto)
        celda.font = fuente_header
        celda.fill = relleno_header
        celda.alignment = Alignment(horizontal="center")
        celda.border = borde

    # ── Datos ──────────────────────────────────────────────────────────
    for fila_idx, item in enumerate(inventario, start=3):
        # Soporte para dict o para objeto Producto
        if isinstance(item, dict):
            datos = [
                item.get('id'),
                item.get('nombre'),
                item.get('stock'),
                item.get('stock_minimo'),
                item.get('precio'),
            ]
            stock_actual  = item.get('stock', 0)
            stock_minimo  = item.get('stock_minimo', 0)
        else:
            datos = [item.id, item.nombre, item.stock, item.stock_minimo, item.precio]
            stock_actual  = item.stock
            stock_minimo  = item.stock_minimo

        es_alerta = stock_actual <= stock_minimo
        color_fila = color_alerta if es_alerta else (color_fila_alt if fila_idx % 2 == 0 else "1e1e2e")

        for col_idx, valor in enumerate(datos, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.font = fuente_normal
            celda.fill = PatternFill("solid", fgColor=color_fila)
            celda.border = borde
            celda.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

        # Columna de estado (6ª)
        estado = ws.cell(row=fila_idx, column=6, value="⚠️ CRÍTICO" if es_alerta else "✅ OK")
        estado.font = Font(name="Calibri", bold=True, size=10,
                           color=(color_alerta if es_alerta else color_ok))
        estado.fill = PatternFill("solid", fgColor=color_fila)
        estado.alignment = Alignment(horizontal="center")

    # Encabezado de la columna estado
    estado_hdr = ws.cell(row=2, column=6, value="Estado")
    estado_hdr.font = fuente_header
    estado_hdr.fill = relleno_header
    estado_hdr.alignment = Alignment(horizontal="center")
    estado_hdr.border = borde

    # ── Anchos de columna ──────────────────────────────────────────────
    anchos = [8, 35, 15, 15, 14, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    wb.save(ruta_salida)
    return ruta_salida
